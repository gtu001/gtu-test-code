
from _decimal import Decimal
from collections import OrderedDict
import csv
from enum import Enum
import inspect
import numbers
from pathlib import Path
import re

import openpyxl
from openpyxl.cell.read_only import EmptyCell
from openpyxl.workbook.workbook import Workbook

from gtu.collection.orderedClass import OrderedClass
from gtu.error import errorHandler
from gtu.io import fileUtil
from gtu.number import numberUtil
from gtu.number import numberUtil
from gtu.openpyxl_test import excelUtil
from gtu.reflect import checkSelf
from gtu.reflect import toStringUtil
from gtu.reflect import toStringUtil
from gtu.string import stringUtil


class EmptyCellException(Exception):
    def __init__(self, message):
        Exception.__init__(self, str(message))

class CellKeeper :
    def __init__(self, cell, rowIndex, colIndex):
        if type(cell) == EmptyCell :
            raise EmptyCellException("此cell 為 empty 需跳過" + str(cell))
#         if cell.has_style:
        self.font = cell.font
        self.border = cell.border
        self.fill = cell.fill
        self.number_format = cell.number_format
        self.protection = cell.protection
        self.alignment = cell.alignment
        self.value = cell.value
        self.rowIndex = rowIndex
        self.colIndex = colIndex
        self.isNumber = numberUtil.isNumber(self.value)

    def apply(self, cell):
        cell.font = self.font
        cell.border = self.border
        cell.fill = self.fill
        cell.number_format = self.number_format
        cell.protection = self.protection
        cell.alignment = self.alignment
        cell.value = self.value
        
    def __repr__(self):
        return toStringUtil.toString(self)
    


def cloneSheetToTargetWb(wb, sheetTitle, cellLst):
    sheet1 = wb.create_sheet(sheetTitle)
    
    for (i, obj) in enumerate(cellLst) :
#         print(i, obj)
        prefix = excelUtil.cellEnglishToPos_toStr(obj.colIndex + 1)
        posStr = prefix + str(obj.rowIndex + 1)
        currentCell = sheet1[posStr]
#         print(posStr, currentCell)
        obj.apply(currentCell)
        
        if i % 100 == 0:
            print("current Index ", i , ", total : " , len(cellLst))
    


def getMaxPos(lst):
    tmpRowIndex = -1;
    tmpColIndex = -1;
    for (i, obj) in enumerate(lst):
        tmpRowIndex = max(obj.rowIndex, tmpRowIndex)
        tmpColIndex = max(obj.colIndex, tmpColIndex)
    return (tmpRowIndex, tmpColIndex)



def getCellLst(sheetTitle, wb, offsetRowCol):
    (offsetRow, offsetCol) = offsetRowCol
    
    lst = list()
    sheet1 = wb[sheetTitle]
    
    for (rowIndex, row) in enumerate(sheet1) :
        for (cellIndex, cell) in enumerate(row) :
            try:
                lst.append(CellKeeper(cell, rowIndex + offsetRow, cellIndex + offsetCol))
            except EmptyCellException as ex :
                print(ex)
            except Exception as ex :
                raise ex
            
    lst.sort(key=lambda x : [x.rowIndex, x.colIndex])
    return lst



def main(excelPath, toExcel):
    
    wb1 = openpyxl.load_workbook(excelPath, read_only=True, data_only=True)
    
    totalLst = list()
    
    tmpRow = 0
    tmpCol = 0
    for (i, name) in enumerate(wb1.get_sheet_names()) :
        
        if name == "年中人口":
            continue
        
        '''
                        只要row遞增即可
        '''
        
        if MERGE_TYPE == 'h' :
            excel1Lst = getCellLst(name, wb1, (0, tmpCol))
        elif MERGE_TYPE == 'v' :
            excel1Lst = getCellLst(name, wb1, (tmpRow, 0))
        else:
            raise Exception("合併方式不可預期(h,v) : " + MERGE_TYPE) 
        
        totalLst.extend(excel1Lst)
        
        (tmpRow, tmpCol) = getMaxPos(totalLst)
        tmpRow += 1
        tmpCol += 1
    
    
    wb = Workbook()
    
    cloneSheetToTargetWb(wb, "結果", totalLst)
    
    wb.save(toExcel)


'''
常數
'''
global MERGE_TYPE #沒這行也行
MERGE_TYPE = 'v' #垂直合併
# MERGE_TYPE = 'h' #水平合併


if __name__ == '__main__':
    excel1 = fileUtil.getCurrentDir() + "RCRP0S108.xlsx"
    toExcel = fileUtil.getDesktopDir() + "活頁簿OK.xlsx"
    main(excel1, toExcel)
    print("done...")
