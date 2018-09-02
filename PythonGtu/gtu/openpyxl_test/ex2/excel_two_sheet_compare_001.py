
from _decimal import Decimal
from collections import OrderedDict
import csv
from enum import Enum
import inspect
import numbers
from pathlib import Path
import re

import openpyxl
from openpyxl.workbook.workbook import Workbook

from gtu.collection.orderedClass import OrderedClass
from gtu.error import errorHandler
from gtu.io import fileUtil
from gtu.number import numberUtil
from gtu.openpyxl_test import excelUtil
from gtu.reflect import checkSelf
from gtu.reflect import toStringUtil
from gtu.string import stringUtil
from gtu.reflect import toStringUtil
from gtu.number import numberUtil



class CellKeeper :
    def __init__(self, cell, rowIndex, colIndex):
#         if cell.has_style:
        self.font = cell.font
        self.border = cell.border
        self.fill = cell.fill
        self.number_format = cell.number_format
        self.protection = cell.protection
        self.alignment = cell.alignment
        self.value = cell.value
        self.rowIndex = rowIndex
        self.colIndex = colIndex
        self.isNumber = numberUtil.isNumber(self.value)

    def apply(self, cell):
        cell.font = self.font
        cell.border = self.border
        cell.fill = self.fill
        cell.number_format = self.number_format
        cell.protection = self.protection
        cell.alignment = self.alignment
        cell.value = self.value
        
    def __repr__(self):
        return toStringUtil.toString(self)
    


def cloneSheetToTargetWb(wb, sheetTitle, cellLst):
    sheet1 = wb.create_sheet(sheetTitle)
    
    for (i, obj) in enumerate(cellLst) :
        print(i, obj)
        prefix = excelUtil.cellEnglishToPos_toStr(obj.colIndex + 1)
        posStr = prefix + str(obj.rowIndex + 1)
        currentCell = sheet1[posStr]
        print(posStr, currentCell)
        obj.apply(currentCell)
    


def getCellLst(excelPath, sheetTitle):
    lst = list()
    wb1 = openpyxl.load_workbook(excelPath, read_only=True, data_only=True)
    sheet1 = wb1[sheetTitle]
    
    for (rowIndex, row) in enumerate(sheet1) :
        for (cellIndex, cell) in enumerate(row) :
            lst.append(CellKeeper(cell, rowIndex, cellIndex))
            
    lst.sort(key=lambda x : [x.rowIndex, x.colIndex])
    return lst



def getFormulaStr(sheetNames, posStr):
    #=工作表1!A1 = 工作表2!A1
    #=IF(原來!J11=新的!J11,TRUE,FALSE)
    formulaStr = "={0}!{2} = {1}!{2}".format(sheetNames[0], sheetNames[1], posStr)
    formulaStr = "=IF({0}!{2} = {1}!{2}, TRUE, FALSE)".format(sheetNames[0], sheetNames[1], posStr)
    return formulaStr



def createCompareSheet(wb, sheetNames, maxPos):
    sheet3 = wb.create_sheet("比對")
    
    rowMax = maxPos[0]
    colMax = maxPos[1]
    
    for row in range(0, rowMax + 1):
        for col in range(0, colMax + 1):
            posStr = excelUtil.cellEnglishToPos_toStr(col + 1) + str(row + 1)
            formula = getFormulaStr(sheetNames, posStr)
            print(formula)
            sheet3[posStr] = formula
            
            

def getMaxPos(lst):
    tmpRowIndex = -1;
    tmpColIndex = -1;
    for (i, obj) in enumerate(lst):
        tmpRowIndex = max(obj.rowIndex, tmpRowIndex)
        tmpColIndex = max(obj.colIndex, tmpColIndex)
    return (tmpRowIndex, tmpColIndex)



def main(excel1, excel2, toExcel):
    excel1Lst = getCellLst(excel1, "工作表1")
    excel2Lst = getCellLst(excel2, "工作表1")
    
    print("excel1Lst", excel1Lst)
    print("excel2Lst", excel2Lst)
            
    wb = Workbook()
    
    cloneSheetToTargetWb(wb, "來自1", excel1Lst)
    cloneSheetToTargetWb(wb, "來自2", excel2Lst)
    
    (maxRow, maxCol) = getMaxPos(excel1Lst);
    
    createCompareSheet(wb, ("來自1", "來自2"), (maxRow, maxCol))
    
    wb.save(toExcel)



if __name__ == '__main__':
    excel1 = fileUtil.getDesktopDir() + "活頁簿1.xlsx"
    excel2 = fileUtil.getDesktopDir() + "活頁簿2.xlsx"
    toExcel = fileUtil.getDesktopDir() + "活頁簿OK.xlsx"
    main(excel1, excel2, toExcel)
    print("done...")
