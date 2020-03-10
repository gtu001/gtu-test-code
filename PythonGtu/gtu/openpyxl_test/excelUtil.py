'''
from gtu.openpyxl_test import excelUtil
'''

import math

import re

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles import Font, Color


def cellEnglishToPos_toInt(column, startByZero=False):
    '''用英文欄位取得相對row的index'''
    '''
columnIndex 從1起算
'''
    column = column.upper().strip()
    total = 0
    _length = len(column)
    for i, c in enumerate(column, 0):
        total += (ord(c) - 64) * pow(26, _length - i - 1)
#     total -= 1
    if startByZero :
        total = total - 1
    return total


def cellEnglishToPos_toStr(columnIndex, startByZero=False):
    '''取回excel的column名'''
    '''
columnIndex 從1起算
'''
    if startByZero :
        columnIndex = columnIndex + 1
    map = {}
    tmpColumn = columnIndex
    while True:
        exponent = 0
        for i in range(1, 1000):
            if math.pow(26, i) >= tmpColumn:
                exponent = i - 1
                break
        tmpColumn = tmpColumn - int(math.pow(26, exponent))
        value = 0
        if map.__contains__(exponent):
            value = map[exponent]
        value += 1
        map[exponent] = value
        if exponent <= 0:
            map[0] = tmpColumn + 1
            break
    rtnVal = ""
    keyLst = list(sorted(map.keys()))
    for i, k in enumerate(keyLst, 0):
        rtnVal = chr(int(map[k]) + 64) + rtnVal
    return rtnVal
    

def debugShowData(sheetName, filePath):
    wb = openpyxl.load_workbook(filePath, read_only=True, data_only=True)
    sheet = wb.get_sheet_by_name(sheetName)
    for rowNum, row in enumerate(sheet, 1):
        str2 = "row" + str(rowNum)
        for cellNum, cell in enumerate(row, 1):
            tmpVal = str(cell.value)
            tmpVal = tmpVal.replace('\n', '')
            tmpVal = tmpVal.replace('\r', '')
            str2 += "\t[" + str(cellNum) + "|" + cellEnglishToPos_toStr(cellNum) + "] " + tmpVal
        print(str2)
        

def getCellValue(column, row, startByZero=False):
    '''
    預設columnIndex 從 1開始 , startByZero=True從0開始
    '''
    offset = -1
    if startByZero :
        offset = 0
    cell = None
    if isinstance(column, int) :
        cell = row[column + offset]
    elif column.isdigit() :
        cell = row[int(column) + offset]
    else :
        columnIdx = cellEnglishToPos_toInt(column, startByZero)
        cell = row[columnIdx] 
    if cell.value is None or \
        str(type(cell.value)) == "<class 'NoneType'>" :
        return ""
    return str(cell.value)


def getSheetByIndex(wb1, index):
    name = wb1.sheetnames[index]
    return wb1[name]


def getRows(sheet, startByZero=False):
    '''
    預設rowIndex 從 1開始 , startByZero=True從0開始
    '''
    if startByZero :
        return list(sheet.rows)
    else :
        lst = list()
        lst.append(None)
        lst.extend(sheet.rows)
        return lst
    
    
def getRowRange(sheet):
    return range(sheet.min_row, sheet.max_row)
    
    
def getCellRange(sheet):
    return range(sheet.min_column, sheet.max_column)


def setCellColor(xls_cell, fgColor=None, bgColor=None):
    '''
        Ex: bgColor = "FFC7CE"
    '''
    if bgColor :
        xls_cell.fill = PatternFill(start_color=bgColor,
                           end_color=bgColor,
                           fill_type='solid')
    if fgColor :
        xls_cell.font = Font(color=fgColor)

    
    
def getCellDefine(cell):
    strVal = str(cell)
    ptn = re.compile("\<.*\'\.(?P<cellDef>\w+)\>", re.I)
    mth = ptn.search(strVal)
    if mth is not None :
        return mth["cellDef"]
    return None

    
if __name__ == '__main__' :
    print(cellEnglishToPos_toStr(3333))
    print("done...")
