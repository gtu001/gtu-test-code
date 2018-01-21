
import csv
from inspect import getmembers

import openpyxl
from openpyxl.workbook.workbook import Workbook

from gtu.io import fileUtil
from gtu.openpyxl_test.ex1.janna.excel_test_rpt_37 import RegionDefEnum, TaoRecac
from gtu.reflect import checkSelf


HAS_G8_TAO = False  # 是否是103年以前(不含)的資料

rowIdMapping = {}

rowIdMapping2 = {}


def getAgeDef(sheet):
    map = {}
    for i, cell in enumerate(sheet[6], 0):
        try:
            map[i] = int(cell.value)
        except TypeError as e :
            pass
        except ValueError as e :
            pass
    
    #R0A ：0歲本年出生者、R0B：0歲去年出生者
    if sheet.title == "36":
        map[4] = "ROA"
        map[5] = "ROB"
        
    return map

class CellDef():
    def __init__(self, value, rowIndex, cellIndex, defMap, sheetName, sheetIndex):
        self.value = value
        self.rowIndex = rowIndex
        self.cellIndex = cellIndex
        self.sheetIndex = sheetIndex
        if self._getRowIdMapping(sheetIndex).__contains__(rowIndex):
            self.region = self._getRowIdMapping(sheetIndex).get(rowIndex)
        else:
            self.region = ""
        self.type = self._getRowIdMapping2(rowIndex)
        if defMap.__contains__(cellIndex) :
            self.age = defMap[cellIndex]
        else:
            self.age = ""
        self.sheet = sheetName
        self.valid = self._isValidValue() 
    
    def _getRowIdMapping(self, sheetIndex):
        return rowIdMapping
        
    def _getRowIdMapping2(self, rowIndex):
        for k, v in rowIdMapping2.items():
            if v.__contains__(rowIndex):
                return k
        return ""
            
    def _isValidValue(self):
        if self.age == '' :
            return False
        if self.type not in ('M', 'F') :
            return False
        if self.region == '':
            return False
        return True
            
    def __repr__(self):
        return "sheet={}, row[{}:{}], val={}, region={}, type={}, age={}, ".format(\
            self.sheet, self.rowIndex, self.cellIndex, self.value, self.region, self.type, self.age) 

def getExcelRow(yyy, row):
    map = {}
    
    f = lambda x: 1 if x == 'M' else 2
    
    map[0] = yyy
    map[1] = "O"
    map[2] = row.region
#     map[3] = f(row.type)
    map[3] = 1 if row.type == 'M' else 2
    map[4] = str(row.age).rjust(3, '0')
    map[5] = row.value
    
    lst = []
    sorted(map)
    for k, v in map.items():
        lst.append(v)
    return lst
    

def writeContent(yyy, lst):
    wb2 = Workbook()
    ws = wb2.active
    ws.append(["STATISTIC_YYY", "STAT_OR", "REGION", "GENDER", "AGE", "DEATH_CNT"])

    for i, row in enumerate(lst, 0):
        if row.valid == True:
            rowArry = getExcelRow(yyy, row)
            ws.append(rowArry)
            
    wb2.save(fileUtil.getDesktopDir() + yyy + "_36.xlsx");
    
    
def writeContentCSV(yyy, lst):
    with open(fileUtil.getDesktopDir() + yyy + "_36.csv", 'w', newline='', encoding='utf-8') as csvfile:
        spamwriter = csv.writer(csvfile, delimiter=',',
                                quotechar='"', quoting=csv.QUOTE_MINIMAL)
        spamwriter.writerow(["STATISTIC_YYY", "STAT_OR", "REGION", "GENDER", "AGE", "DEATH_CNT"])
        for i, row in enumerate(lst, 0):
            if row.valid == True:
                rowArry = getExcelRow(yyy, row)
                spamwriter.writerow(rowArry)


def main(filePath, yyy):
    lst = []
    wb = openpyxl.load_workbook(filePath, 'r', data_only=True)
    for shIndex, name in enumerate(wb.sheetnames, 0):
        sheet = wb.get_sheet_by_name(name)
        
        regionMap = RegionDefEnum.getRegionDefMap(sheet, -1)#offset 記得調整 XXX
        if len(regionMap) != 0:
            global rowIdMapping
            rowIdMapping = regionMap
            global rowIdMapping2
            rowIdMapping2 = RegionDefEnum.getTMFDefMap(regionMap)
        
        defMap = getAgeDef(sheet);
        
        for i, row in enumerate(sheet, 1):
            for j, cell in enumerate(row, 0):
                newVal = CellDef(cell.value, i, j, defMap, name, shIndex)
                lst.append(newVal)
                
    # 桃園重算    
    if HAS_G8_TAO :
        TaoRecac(lst).caclute()
            
    # writeContent(yyy, lst)
    writeContentCSV(yyy, lst)



if __name__ == '__main__' :
    main("C:/Users/gtu001/Desktop/101-105/101年刊/101-36_Janna.xlsx", "101")
    print("done...")
        
        
