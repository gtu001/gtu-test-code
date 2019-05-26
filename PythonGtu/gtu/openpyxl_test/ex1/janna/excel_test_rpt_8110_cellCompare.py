
'''
from gtu.openpyxl_test.ex1.janna import excel_test_rpt_8110_cellCompare
'''

from _decimal import Decimal
from enum import Enum
import inspect
import numbers
import os
from pathlib import Path
import re

import openpyxl
from openpyxl.workbook.workbook import Workbook

from gtu._tkinter.util.tkinterUIUtil import _ProgressBar
from gtu.collection.orderedClass import OrderedClass
from gtu.error import errorHandler
from gtu.io import fileUtil
from gtu.number import numberUtil
from gtu.openpyxl_test import excelUtil
from gtu.reflect import checkSelf
from gtu.regex import regexReplace
from gtu.string import stringUtil
from gtu.util import rangeUtil


def getRangeWarpper(val1, val2):
    start = min(val1, val2)
    end = max(val1, val2)
    return rangeUtil.RangeWrapper(range(start, end))


def fixCellValue_None(strVal):
    if strVal is None :
        return ""
    return strVal


def fixCellValue_Replace(strVal):
    if strVal is None :
        return ""
    strVal = strVal.replace(',', '').strip()
    return strVal


def fixCellValue_Zero(strVal):
    if strVal is None :
        return ""
    if strVal == '-' :
        return ""
    if stringUtil.isNumber(strVal) and \
        float(strVal) == 0:
        return ""
    return strVal
    
    
def isNumberEqual(val1, val2):
    try:
        if float(val1) == float(val2) :
            return True
        else :
            return False
    except :
        return None


class MyProgressHander():
    def __init__(self, mProgressBar):
        self.mProgressBar = mProgressBar
    
    def validateOk(self):
        return self.mProgressBar is not None and \
            isinstance(self.mProgressBar, _ProgressBar)
    
    def setupValue(self, initVal, maxVal):
        if self.validateOk():
            self.mProgressBar.setupValue(initVal, maxVal)
            
    def step(self, amount=None):
        if self.validateOk():
            print("step ", amount)
            self.mProgressBar.step(amount)


class LogSb():
    def __init__(self, logSb):
        self.logSb = logSb
        self.isOk = logSb is not None
        
    def println(self, line):
        if self.isOk :
            self.logSb.appendText(line + "\n")

        
def mainDetail(fileC, fileI, mProgressBar=None, logSb=None):
    
    mPBar = MyProgressHander(mProgressBar)
    
    global logger
    logger = LogSb(logSb)
    
    diffLog = _DiffLog()
    
    wb1 = openpyxl.load_workbook(fileC, read_only=False, data_only=True)
    wb2 = openpyxl.load_workbook(fileI, read_only=False, data_only=True)
    
    
    maxSheetRng = getRangeWarpper(len(wb1.get_sheet_names()), len(wb2.get_sheet_names()))
    
    #設定進度條基本數值
    mPBar.setupValue(0, maxSheetRng.left)
    
    sameCount = 0
    diffCount = 0
    
    for sh_idx in range(0, maxSheetRng.left) :
        s1Name = wb1.get_sheet_names()[sh_idx];
        s2Name = wb2.get_sheet_names()[sh_idx];
        
        print("開始比較 : ", sh_idx, s1Name, s2Name)
        
        sheet1 = wb1.get_sheet_by_name(s1Name)
        sheet2 = wb2.get_sheet_by_name(s2Name)
        
        if diffLog.templateDefineLst is None :
            diffLog.initTemplateDefineLst(sheet1)
            
        
        maxRowRng = getRangeWarpper(sheet1.max_row, sheet2.max_row)
        maxColRng = getRangeWarpper(sheet1.max_column, sheet2.max_column)
        
        print("rowLeft ", maxRowRng.left)
        
        for row_idx in range(1, maxRowRng.left) :  # maxRowRng.left <-- RCRP0C210_C8_1 最多到 row:43
            row_idx_1 = row_idx
            row_idx_2 = row_idx
            
            diff = _ColDiff.findRowFrom(row_idx)
            if diff :
                row_idx_2 = diff.rowTo
            
                    
            row1 = excelUtil.getRows(sheet1)[row_idx_1]
            row2 = excelUtil.getRows(sheet2)[row_idx_2]
            
            for col_idx in range(1, maxColRng.left) :
#                 print("--------------------------------------------")
#                 print(excelUtil.getCellDefine(row1[col_idx -1]))
#                 print(excelUtil.getCellDefine(row2[col_idx -1]))
                col_idx_1 = col_idx
                col_idx_2 = col_idx

                if diff :
                    col_idx_2 = col_idx + diff.colDiff
                
                cell1Val = excelUtil.getCellValue(col_idx_1, row1)
                cell2Val = excelUtil.getCellValue(col_idx_2, row2)
                cell1Val = fixCellValue_None(cell1Val)
                cell2Val = fixCellValue_None(cell2Val)
                cell1Val = stringUtil.trimSpace(cell1Val)
                cell2Val = stringUtil.trimSpace(cell2Val)
                cell1Val = fixCellValue_Zero(cell1Val)
                cell2Val = fixCellValue_Zero(cell2Val)
                cell1Val = fixCellValue_Replace(cell1Val)
                cell2Val = fixCellValue_Replace(cell2Val)
                
                cellEng1 = excelUtil.cellEnglishToPos_toStr(col_idx_1) + str(row_idx_1)
                cellEng2 = excelUtil.cellEnglishToPos_toStr(col_idx_2) + str(row_idx_2)
                
                diffLogBean = _DiffLogBean(s1Name + "," + s2Name, \
                                           cellEng1, \
                                           cellEng2, \
                                           cell1Val, \
                                           cell2Val \
                                           )
                
                numberCompareResult = isNumberEqual(cell1Val, cell2Val)
                
                if numberCompareResult is not None and numberCompareResult == False : 
                    diffLog.add(diffLogBean)
                    diffCount += 1
                else :
                    sameCount += 1
        
        #完成一個表
        mPBar.step(1)
        
    diffLog.writeExcel()
    
    print("sameCount", sameCount)
    print("diffCount", diffCount)
                
#         openpyxl.worksheet.worksheet.Worksheet



class _DiffLogBean():

    def __init__(self, sheetName, cellEng1, cellEng2, cellValue1, cellValue2):
        self.sheetName = sheetName
        self.cellEng1 = cellEng1
        self.cellEng2 = cellEng2
        self.cellValue1 = cellValue1
        self.cellValue2 = cellValue2


class _DiffLog():

    def __init__(self):
        self.sheetNameCellMap = dict()
        self.templateDefineLst = None
    
    def add(self, diffLogBean):
        lst = list()
        if self.sheetNameCellMap.__contains__(diffLogBean.sheetName) :
            lst = self.sheetNameCellMap[diffLogBean.sheetName]
        lst.append(diffLogBean)
        self.sheetNameCellMap[diffLogBean.sheetName] = lst
        
    def writeExcel(self):
        wb = Workbook()
        
        sheetIdx = 0
        
        for (i, key) in enumerate(self.sheetNameCellMap.keys()) :
            sheet1 = wb.create_sheet(key, sheetIdx)
            
            self.writeTemplateDefineLstCell(sheet1)
            
            lst = self.sheetNameCellMap[key]
            for (j, bean) in enumerate(lst) :
                print("Diff cell : ", bean.cellEng1, bean.cellEng2, bean.cellValue1, bean.cellValue2)
                sheet1[bean.cellEng1].value = "[C]" + bean.cellValue1 + " <-> " + "[I]" + bean.cellValue2
                excelUtil.setCellColor(sheet1[bean.cellEng1], bgColor="FDE6E0")
                logger.println("頁簽 :" + key + " 欄位不同 , [C]" + bean.cellEng1 + " = " + bean.cellValue1 + ", [I]" + bean.cellEng2 + " = " + bean.cellValue2)
        
            sheetIdx += 1
            
        logFile = fileUtil.getDesktopDir() + os.sep + "diff_RCRP0C210_C8_1.xlsx"
        print("log file ", logFile)
        
        wb.save(logFile)
        
        
    def initTemplateDefineLst(self, sheet1):
        templateDict = dict()
        titleRowIndexLst = [5, 6, 7, 10, 11, 12, 15, 16, 17, 20, 21, 22 , 25, 26, 29, 30, 33, 34, 35, 38, 39]
        for (i1, row_idx) in enumerate(titleRowIndexLst) :
            for col_idx in range(1, sheet1.max_column) :
                cellEng1 = excelUtil.cellEnglishToPos_toStr(col_idx) + str(row_idx)
                templateDict[cellEng1] = sheet1[cellEng1].value
        self.templateDefineLst = templateDict
        
        
    def writeTemplateDefineLstCell(self, sheet1):
        for (i, key) in enumerate(self.templateDefineLst.keys()) :
            value = self.templateDefineLst[key]
            sheet1[key] = value
        

class _ColDiff(Enum):
    B8 = ("B8", 8, 8, 0)
    B9 = ("B9", 9, 9, 0)
    B13 = ("B13", 13, 13, 0)
    B14 = ("B14", 14, 14, 0)
    B18 = ("B18", 18, 18, 0)
    B19 = ("B19", 19, 19, 0)
    B23 = ("B23", 23, 23, 0)
    B24 = ("B24", 24, 24, 0)
    C25 = ("B25", 25, 25, -1)
    C26 = ("B26", 26, 26, -1)
    C27 = ("B28", 27, 28, -1)
    C28 = ("B29", 28, 29, -1)
    C29 = ("B30", 29, 30, -1)
    C30 = ("B31", 30, 31, -1)
    C31 = ("B33", 31, 33, -1)
    C32 = ("B34", 32, 34, -1)
    C33 = ("B35", 33, 35, -1)
    C34 = ("B36", 34, 36, -1)
    C35 = ("B37", 35, 37, -1)
    C36 = ("B38", 36, 38, -1)
    C37 = ("B39", 37, 39, -1)
    C38 = ("C41", 38, 41, 0)
    C39 = ("C42", 39, 42, 0)
    C40 = ("C43", 40, 43, 0)
    C41 = ("C44", 41, 44, 0)
    C42 = ("C45", 42, 45, 0)
    C43 = ("C46", 43, 46, 0)
    
    def __init__(self, startLabel, rowFrom, rowTo, colDiff):
        self.startLabel = startLabel
        self.rowFrom = rowFrom
        self.rowTo = rowTo
        self.colDiff = colDiff
        
    @classmethod
    def findRowFrom(clz, rowFrom):  
        for i, name in enumerate(_ColDiff.__members__, 0):
            e = _ColDiff[name]
            if e.rowFrom == rowFrom :
                return e
        return None    
    

if __name__ == '__main__':
#     fileC = fileUtil.getDesktopDir() + os.sep + "/janna/善良瓜新工作/RCRP0C210_C8_1.XLSX"
#     fileI = fileUtil.getDesktopDir() + os.sep + "/janna/善良瓜新工作/RCRP0C210_I8_1.XLSX"
    
    fileC = 'C:/Users/wistronits/Desktop/新增資料夾/TX0CRC1902181339357041_RCRP0C210_C2.xlsx'
    fileI = 'C:/Users/wistronits/Desktop/新增資料夾/TX0CRC1902181339357041_RCRP0C210_I2.xlsx'

    mainDetail(fileC, fileI)
    print("done...")
    
    
    
    
    
    
    
    
    
    
    
    
    
    
