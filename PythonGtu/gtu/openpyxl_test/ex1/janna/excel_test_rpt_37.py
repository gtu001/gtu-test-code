
import csv
from enum import Enum
from inspect import getmembers
import re

import openpyxl
from openpyxl.workbook.workbook import Workbook

from gtu.base import equalUtil
from gtu.collection import mapUtil
from gtu.io import fileUtil
from gtu.reflect import checkSelf
from gtu.regex import regexUtil
from gtu.string import stringUtil


HAS_G8_TAO = False  # 是否是103年以前(不含)的資料

rowIdMapping = {}

rowIdMapping2 = {}

class AgeTypeDef(Enum):
    R05 = ("5～9")
    R10 = ("10～14")
    R15 = ("15～19")
    R20 = ("20～24")
    R25 = ("25～29")
    R30 = ("30～34")
    R35 = ("35～39")
    R40 = ("40～44")
    R45 = ("45～49")
    R50 = ("50～54")
    R55 = ("55～59")
    R60 = ("60～64")
    R65 = ("65～69")
    R70 = ("70～74")
    R75 = ("75～79")
    R80 = ("80～84")
    R85 = ("85～89")
    R90 = ("90～94")
    R95 = ("95～99")
    _100 = ("100+")
    
    def __init__(self, strVal):
        self.strVal = strVal
        
    @staticmethod
    def checkAgeType(strVal):
        for i, name in enumerate(AgeTypeDef.__members__, 0):
            e = AgeTypeDef[name]
            if e.strVal == strVal:
                return e
        return None
            
def getAgeDef(sheet):
    map = {}
    for i, cell in enumerate(sheet[4], 0):
        type = AgeTypeDef.checkAgeType(cell.value)
        if type != None:
            tmpVal = type.name
            if tmpVal.startswith('_'):
                tmpVal = tmpVal.replace('_', '')
            map[i] = tmpVal
    
    # R0A ：0歲本年出生者、R0B：0歲去年出生者
    if sheet.title in ("37", "37-一"):
        map[4] = "R0A"
        map[5] = "R0B"
        map[7] = "001"
        map[8] = "002"
        map[9] = "003"
        map[10] = "004"
        
    print(map)
        
    return map

class CellDef():
    def __init__(self, value, rowIndex, cellIndex, defMap, sheetName, sheetIndex):
        self.value = value
        self.rowIndex = rowIndex
        self.cellIndex = cellIndex
        self.sheetIndex = sheetIndex
        if self._getRowIdMapping(sheetIndex).__contains__(rowIndex):
            self.region = self._getRowIdMapping(sheetIndex).get(rowIndex)
        else:
            self.region = ""
        self.type = self._getRowIdMapping2(rowIndex)
        if defMap.__contains__(cellIndex) :
            self.age = defMap[cellIndex]
        else:
            self.age = ""
        self.sheet = sheetName
        self.valid = self._isValidValue() 
        self.hashDef = equalUtil.hashInclude(self, ["valid", "type", "age", "region"])
    
    def _getRowIdMapping(self, sheetIndex):
        return rowIdMapping
        
    def _getRowIdMapping2(self, rowIndex):
        for k, v in rowIdMapping2.items():
            if v.__contains__(rowIndex):
                return k
        return ""
            
    def _isValidValue(self):
        if self.age == '' :
            return False
        if self.type not in ('M', 'F') :
            return False
        if self.region == '':
            return False
        if not str(self.value).isdigit() :
            raise Exception("ERROR : " + self.__repr__())
        return True
            
    def __repr__(self):
        return "sheet={}, row[{}:{}], val={}, region={}, type={}, age={}, ".format(\
            self.sheet, self.rowIndex, self.cellIndex, self.value, self.region, self.type, self.age) 

def getExcelRow(yyy, row):
    map = {}
    
    f = lambda x: 1 if x == 'M' else 2
    
    map[0] = yyy
    map[1] = "O"
    map[2] = row.region
#     map[3] = f(row.type)
    map[3] = 1 if row.type == 'M' else (2 if row.type == 'F' else '')
    map[4] = str(row.age).rjust(3, '0')
    map[5] = row.value
    
    lst = []
    sorted(map)
    for k, v in map.items():
        lst.append(v)
    return lst
    

def writeContent(yyy, lst):
    wb2 = Workbook()
    ws = wb2.active
    ws.append(["STATISTIC_YYY", "STAT_OR", "REGION", "GENDER", "AGE", "DEATH_CNT"])

    for i, row in enumerate(lst, 0):
        if row.valid == True:
            rowArry = getExcelRow(yyy, row)
            ws.append(rowArry)
            
    wb2.save(fileUtil.getDesktopDir() + yyy + "_37.xlsx");
    
    
def writeContentCSV(yyy, lst):
    with open(fileUtil.getDesktopDir() + yyy + "_37_new.csv", 'w', newline='', encoding='utf-8') as csvfile:
        spamwriter = csv.writer(csvfile, delimiter=',',
                                quotechar='"', quoting=csv.QUOTE_MINIMAL)
        spamwriter.writerow(["STATISTIC_YYY", "STAT_OR", "REGION", "GENDER", "AGE", "DEATH_CNT"])
        for i, row in enumerate(lst, 0):
            if row.valid == True:
                rowArry = getExcelRow(yyy, row)
                spamwriter.writerow(rowArry)


def main(filePath, yyy):
    lst = []
    wb = openpyxl.load_workbook(filePath, 'r', data_only=True)
#     for shIndex, name in enumerate(wb.sheetnames, 0):
    sheetNames = ["37", "37-0", "37-一", "37-1"]
    for shIndex, name in enumerate(sheetNames, 0):
        sheet = wb.get_sheet_by_name(name)
        print("sheet - ", name)
        
        regionMap = RegionDefEnum.getRegionDefMap(sheet, 0)
        if len(regionMap) != 0:
            global rowIdMapping
            rowIdMapping = regionMap
            global rowIdMapping2
            rowIdMapping2 = RegionDefEnum.getTMFDefMap(regionMap)
        
        defMap = getAgeDef(sheet);
        
        for i, row in enumerate(sheet, 1):
            for j, cell in enumerate(row, 0):
                newVal = CellDef(cell.value, i, j, defMap, name, shIndex)
                addIfDuplicate(lst, newVal)
                
    # 桃園重算    
    if HAS_G8_TAO :
#         TaoRecac(lst).caclute()
        pass
            
    # writeContent(yyy, lst)
    writeContentCSV(yyy, lst)
    

def addIfDuplicate(lst, single):
    isAdd = False
    for i, row in enumerate(lst, 0):
        if row.hashDef == single.hashDef and \
            equalUtil.equalsInclude(row, single, ["type", "age", "region"]) and \
            str(row.value).isdigit() and \
            str(single.value).isdigit() and \
            row.valid and single.valid :
            print("重複", row, single)
            row.value = str(int(row.value) + int(single.value))
            isAdd = True
    if isAdd == False:
        lst.append(single)
    
    
class TaoRecac():
    def __init__(self, lst):
        self.lst = lst
        
    def caclute(self):
        for row in self.lst:
            if row.valid and row.region == '10000000':
                newVal = self.taoRecaclute2(row)
                row.value = str(newVal)
                
    def taoRecaclute2(self, row):
        total = 0
        for r in self.lst:
            if r.valid and r.type == row.type and r.age == row.age and \
                not r.region.startswith("100") :
                total += int(r.value)
        return total
    
    
class RegionDefEnum(Enum):
    Region1 = ("總計")
    RegionZ1 = ("臺閩地區")
    Region65000000 = ("新北市")
    Region63000000 = ("臺北市")
    Region68000000 = ("桃園市")
    RegionA68000000 = ("桃園縣")
    Region66000000 = ("臺中市")
    Region67000000 = ("臺南市")
    Region64000000 = ("高雄市")
    Region10000000 = ("臺灣省")
    Region10002000 = ("宜蘭縣")
    Region10004000 = ("新竹縣")
    Region10005000 = ("苗栗縣")
    Region10007000 = ("彰化縣")
    Region10008000 = ("南投縣")
    Region10009000 = ("雲林縣")
    Region10010000 = ("嘉義縣")
    Region10013000 = ("屏東縣")
    Region10014000 = ("臺東縣")
    Region10015000 = ("花蓮縣")
    Region10016000 = ("澎湖縣")
    Region10017000 = ("基隆市")
    Region10018000 = ("新竹市")
    Region10020000 = ("嘉義市")
    Region09 = ("福建省")
    Region09020000 = ("金門縣")
    Region09007000 = ("連江縣")
    #舊資料
    RegionOld65000000=("臺北縣")
    RegionOld10002000=("宜蘭縣")
    RegionOld68000000=("桃園縣")
    RegionOld10004000=("新竹縣")
    RegionOld10005000=("苗栗縣")
    RegionOldZ66000000=("臺中縣")
    RegionOld10007000=("彰化縣")
    RegionOld10008000=("南投縣")
    RegionOld10009000=("雲林縣")
    RegionOld10010000=("嘉義縣")
    RegionOldZ67000000=("臺南縣")
    RegionOldZ64000000=("高雄縣")
    RegionOld10013000=("屏東縣")
    RegionOld10014000=("臺東縣")
    RegionOld10015000=("花蓮縣")
    RegionOld10016000=("澎湖縣")
    RegionOld10017000=("基隆市")
    RegionOld10018000=("新竹市")
    RegionOld66000000=("臺中市")
    RegionOld10020000=("嘉義市")
    RegionOld67000000=("臺南市")
    RegionOld63000000=("臺北市")
    RegionOld64000000=("高雄市")
    RegionOld09=("福建省")
    RegionOld09020000=("金門縣")
    RegionOld09007000=("連江縣")

    def __init__(self, chs):
        self.chs = chs
    
    @staticmethod
    def findRegion(chs):
        for i, name in enumerate(RegionDefEnum.__members__, 0):
            e = RegionDefEnum[name]
            if e.chs == chs :
                line = e.name
                line = re.sub(r"[a-zA-Z]", "", line)
                return line
        return None
    
    @staticmethod
    def getRegionDefMap(sheet, offset):
        map = {}
        for i, row in enumerate(sheet, 1):
            strVal = str(row[0].value)
            strVal = strVal.replace('\\s', '').replace('　', '').replace(' ', '').strip()
            if stringUtil.isChinese(strVal):
                region = RegionDefEnum.findRegion(strVal)
                if region != None :
                    print("-->", i, strVal, region)
                    map[i + offset] = region
                    map[i + offset + 1] = region
                    map[i + offset + 2] = region
                    if strVal == "桃園縣": 
                        global HAS_G8_TAO
                        HAS_G8_TAO = True
                else:
#                     raise Exception("region NotFound : ", strVal)
                    print("region NotFound : ", strVal)
        print("regionMap = ", map)
        return map
    
    @staticmethod
    def getTMFDefMap(regionMap):
        map = {}
        st = set(regionMap.keys())
        sorted(st)
        tmpCount = 0
        tmpVal = "";
        for k in st:
            if tmpVal != regionMap[k]:
                tmpCount = 0
                tmpVal = regionMap[k]
                mapUtil.addMap('T', k, map)
            else:
                tmpCount += 1
                if tmpCount == 1:
                    mapUtil.addMap('M', k, map)
                else:
                    mapUtil.addMap('F', k, map)
        print("TMF defMap = ", map)
        return map
    
if __name__ == '__main__' :
#     main("C:/Users/gtu001/Desktop/db_data_歷年/101-105/101年刊/101-37.xlsx", "101")
#     main("C:/Users/gtu001/Desktop/db_data_歷年/96-100/100-37.xlsx", "100")
#     main("C:/Users/gtu001/Desktop/db_data_歷年/101-105/102年刊/102-37.xlsx", "102")
#     main("C:/Users/gtu001/Desktop/db_data_歷年/101-105/103年刊/103-37.xlsx", "103")
#     main("C:/Users/gtu001/Desktop/db_data_歷年/101-105/104年刊/104-37.xlsx", "104")
#     main("C:/Users/gtu001/Desktop/db_data_歷年/101-105/105_年刊/105-37_janna.xlsx", "105")

#     main("C:/Users/gtu001/Desktop/db_data_歷年/96-100/96-37.xlsx", "96") 
    main("C:/Users/gtu001/Desktop/db_data_歷年/96-100/97-37.xlsx", "97")
    main("C:/Users/gtu001/Desktop/db_data_歷年/96-100/98-37.xlsx", "98")
    main("C:/Users/gtu001/Desktop/db_data_歷年/96-100/99-37.xlsx", "99")

#     main("C:/SA/綠色便民案/06.開發文件/db_data/96-100/96-37.xlsx", "96")

    print("done...")
        
        
