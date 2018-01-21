
import csv
from enum import Enum
from inspect import getmembers
import re

import openpyxl
from openpyxl.workbook.workbook import Workbook

from gtu.io import fileUtil
from gtu.openpyxl_test.ex1.janna.excel_test_rpt_37 import RegionDefEnum, TaoRecac
from gtu.reflect import checkSelf
from gtu.string import stringUtil


HAS_G8_TAO = False  # 是否是103年以前(不含)的資料

rowIdMapping = {}

rowIdMapping2 = {}


def getDistinctRegion():
    lst = set()
    for i, name in enumerate(RegionDefEnum.__members__, 0) :
        val = re.sub(r"[a-zA-Z]", "", name)
        lst.add(val)
    return lst


def getAgeDef(sheet):
    map = {}
    for i, cell in enumerate(sheet[5], 0):
        try:
            strVal = str(cell.value)
            if strVal.__contains__("100"):
                strVal = "100"
            map[i] = int(strVal)
        except TypeError as e :
            pass
        except ValueError as e :
            pass
    return map


class CellDef():    
    def __init__(self, value, rowIndex, cellIndex, defMap, sheetName, sheetIndex):
        self.value = value
        self.rowIndex = rowIndex
        self.cellIndex = cellIndex
        self.sheetIndex = sheetIndex
        if self._getRowIdMapping(sheetIndex).__contains__(rowIndex):
            self.region = self._getRowIdMapping(sheetIndex).get(rowIndex)
        else:
            self.region = ""
        self.type = self._getRowIdMapping2(rowIndex)
        if defMap.__contains__(cellIndex) :
            self.age = defMap[cellIndex]
        else:
            self.age = ""
        self.sheet = sheetName
        self.valid = self._isValidValue() 
    
    def _getRowIdMapping(self, sheetIndex):
        return rowIdMapping
        
    def _getRowIdMapping2(self, rowIndex):
        for k, v in rowIdMapping2.items():
            if v.__contains__(rowIndex):
                return k
        return ""
            
    def _isValidValue(self):
        if self.age == '' :
            return False
        if self.type not in ('M', 'F') :
            return False
        if self.region == '':
            return False
        return True
            
    def __repr__(self):
        return "sheet={}, row[{}:{}], val={}, region={}, type={}, age={}, ".format(\
            self.sheet, self.rowIndex, self.cellIndex, self.value, self.region, self.type, self.age) 


class CellDef2(CellDef):
    def __init__(self, age, type, region, value):
        if age.startswith('_') :
            age = age.replace('_', '')
        self.age = age
        self.type = type
        self.region = region
        self.value = value
        self.valid = True
        self._fakeData()
        
    def _fakeData(self):
        self.sheet = ""
        self.cellIndex = -1
        self.rowIndex = -1
        
    def __repr__(self):
        return "sheet={}, row[{}:{}], val={}, region={}, type={}, age={}, ".format(\
            "NA", "NA", "NA", self.value, self.region, self.type, self.age) 


def getExcelRow(yyy, row):
    map = {}
    
    f = lambda x: 1 if x == 'M' else 2
    
    map[0] = yyy
    map[1] = row.region
    map[2] = row.region  # ADMIN_OFFICE_CODE
#     map[3] = f(row.type)
    map[3] = row.region  # SITE_ID
    map[4] = 1 if row.type == 'M' else 2
    
    if isinstance(row, CellDef2) :
        map[5] = "5" #子類別要優先判斷否則會無法辨識
    elif isinstance(row, CellDef) :
        map[5] = "1"
    
    map[6] = str(row.age).rjust(3, '0')
    map[7] = row.value
    
    lst = []
    sorted(map)
    for k, v in map.items():
        lst.append(v)
    return lst
    

def writeContent(yyy, lst):
    wb2 = Workbook()
    ws = wb2.active
    ws.append(["STATISTIC_YYY", "REGION", "ADMIN_OFFICE_CODE", "SITE_ID", "GENDER", "AGE_TYPE", "AGE", "CNT"])
    for i, row in enumerate(lst, 0):
        if row.valid == True:
            rowArry = getExcelRow(yyy, row)
            ws.append(rowArry)
    wb2.save(fileUtil.getDesktopDir() + yyy + "_4.xlsx");
    
    
def writeContentCSV(yyy, lst):
    with open(fileUtil.getDesktopDir() + yyy + "_4_一年組.csv", 'w', newline='', encoding='utf-8') as csvfile:
        spamwriter = csv.writer(csvfile, delimiter=',',
                                quotechar='"', quoting=csv.QUOTE_MINIMAL)
        spamwriter.writerow(["STATISTIC_YYY", "REGION", "ADMIN_OFFICE_CODE", "SITE_ID", "GENDER", "AGE_TYPE", "AGE", "CNT"])
        for i, row in enumerate(lst, 0):
            if row.valid == True:
                rowArry = getExcelRow(yyy, row)
                spamwriter.writerow(rowArry)
                

def writeContentCSV2(yyy, lst):
    with open(fileUtil.getDesktopDir() + yyy + "_4_五年組.csv", 'w', newline='', encoding='utf-8') as csvfile:
        spamwriter = csv.writer(csvfile, delimiter=',',
                                quotechar='"', quoting=csv.QUOTE_MINIMAL)
        spamwriter.writerow(["STATISTIC_YYY", "REGION", "ADMIN_OFFICE_CODE", "SITE_ID", "GENDER", "AGE_TYPE", "AGE", "CNT"])
        for i, row in enumerate(lst, 0):
            if row.valid == True and type(row) == CellDef2:
                rowArry = getExcelRow(yyy, row)
                spamwriter.writerow(rowArry)


def caculateAgeGroup(lst):
    newLst = []
    sexLst = ['M', 'F']
    regionLst = getDistinctRegion()
    for i, region in enumerate(regionLst, 0):
        for j, sex in enumerate(sexLst, 0):
            for i, name in enumerate(AgeGroup.__members__, 0):
                e = AgeGroup[name]
#                 print(i, e.name, e.fromAge, e.toAge)
                newVal = e.sumRegionValue(region, sex, lst)
                newLst.append(CellDef2(e.name, sex, region, newVal))
    return newLst


class AgeGroup(Enum):
    _0 = (0, 0)
    _1 = (1, 1)
    _2 = (2, 2)
    _3 = (3, 3)
    _4 = (4, 4)
#     R01 = (1, 4)
    R05 = (5, 9)
    R10 = (10, 14)
    R15 = (15, 19)
    R20 = (20, 24)
    R25 = (25, 29)
    R30 = (30, 34)
    R35 = (35, 39)
    R40 = (40, 44)
    R45 = (45, 49)
    R50 = (50, 54)
    R55 = (55, 59)
    R60 = (60, 64)
    R65 = (65, 69)
    R70 = (70, 74)
    R75 = (75, 79)
    R80 = (80, 84)
    R85 = (85, 89)
    R90 = (90, 94)
    R95 = (95, 99)
    _100 = (100, 130)
    
    def __init__(self, fromAge, toAge):
        self.fromAge = fromAge
        self.toAge = toAge
        
    def sumRegionValue(self, region, sex, lst):
        total = 0
        for i, row in enumerate(lst, 0):
            if row.valid == True :
                _age = int(row.age) 
                if _age >= self.fromAge and _age <= self.toAge and \
                    row.region == region and \
                    row.type == sex :
                    total += int(row.value)
        return total


def main(filePath, yyy):
    lst = []
    wb = openpyxl.load_workbook(filePath, 'r', data_only=True)
    for shIndex, name in enumerate(wb.sheetnames, 0):
        sheet = wb.get_sheet_by_name(name)
        
        regionMap = RegionDefEnum.getRegionDefMap(sheet, -1)
        if len(regionMap) != 0:
            global rowIdMapping
            rowIdMapping = regionMap
            global rowIdMapping2
            rowIdMapping2 = RegionDefEnum.getTMFDefMap(regionMap)
        
        defMap = getAgeDef(sheet);
        print("ageDef", defMap)
        
        for i, row in enumerate(sheet, 1):
            for j, cell in enumerate(row, 0):
                newVal = CellDef(cell.value, i, j, defMap, name, shIndex)
                lst.append(newVal)
                
                
    newLst = caculateAgeGroup(lst)
        
    # 桃園重算    
    if HAS_G8_TAO :
        TaoRecac(lst).caclute()
        
    #duplicateCheck(lst)
            
    # writeContent(yyy, lst)
    writeContentCSV(yyy, lst)
    
    # writeContent(yyy, lst)
    writeContentCSV2(yyy, newLst)
    
def duplicateCheck(lst):
    lst2 = []
    for i,row1 in enumerate(lst, 0):
        if isinstance(row1, CellDef2):
            lst2.append(row1)
    
    for i,row1 in enumerate(lst2, 0):
        for j,row2 in enumerate(lst2, 0):
            if i != j and \
                row1.age == row2.age and \
                row1.type == row2.type and \
                row1.region == row2.region and \
                row1.valid and row2.valid :
                print("ERROR ", i, j, row1, row2)

if __name__ == '__main__' :
#     main("C:/Users/gtu001/Desktop/db_data_歷年/101-105/101年刊/101-04.xlsx", "101")
#     main("C:/Users/gtu001/Desktop/db_data_歷年/101-105/102年刊/102-04.xlsx", "102")
#     main("C:/Users/gtu001/Desktop/db_data_歷年/101-105/103年刊/103-04.xlsx", "103")
#     main("C:/Users/gtu001/Desktop/db_data_歷年/101-105/104年刊/104-04.xlsx", "104")
#     main("C:/Users/gtu001/Desktop/db_data_歷年/101-105/105_年刊/105-04.xlsx", "105")
    main("C:/SA/綠色便民案/06.開發文件/db_data/報表04/105-04.xlsx", "105")

    print("done...")
        
        
