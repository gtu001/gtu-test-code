
from _decimal import Decimal
from collections import OrderedDict
import csv
from enum import Enum
import inspect
import numbers
import os
from pathlib import Path
import re

import openpyxl
from openpyxl.workbook.workbook import Workbook

from gtu.collection.orderedClass import OrderedClass
from gtu.error import errorHandler
from gtu.io import fileUtil
from gtu.number import numberUtil
from gtu.openpyxl_test import excelUtil
from gtu.reflect import checkSelf
from gtu.regex import regexReplace
from gtu.string import stringUtil
from gtu.util import rangeUtil


def getRangeWarpper(val1, val2):
    start = min(val1, val2)
    end = max(val1, val2)
    return rangeUtil.RangeWrapper(range(start, end))


def fixCellValue_None(strVal):
    if strVal is None :
        return ""
    return strVal


def fixCellValue_Replace(strVal):
    if strVal is None :
        return ""
    strVal = strVal.replace(',', '').strip()
    return strVal


def fixCellValue_Zero(strVal):
    if strVal is None :
        return ""
    if strVal == '-' :
        return ""
    if stringUtil.isNumber(strVal) and \
        float(strVal) == 0:
        return ""
    return strVal
    
    
    
def isNumberEqual(val1, val2):
    try:
        if float(val1) == float(val2) :
            return True
    except :
        return False
    
    
        
def mainDetail(file1, file2):
    wb1 = openpyxl.load_workbook(file1, read_only=True, data_only=True)
    wb2 = openpyxl.load_workbook(file2, read_only=True, data_only=True)
    wb3 = Workbook()
    
    maxSheetRng = getRangeWarpper(len(wb1.get_sheet_names()), len(wb2.get_sheet_names()))
    
    for sh_idx in range(0, maxSheetRng.left) :
        s1Name = wb1.get_sheet_names()[sh_idx];
        s2Name = wb2.get_sheet_names()[sh_idx];
        
        print("開始比較 : ", sh_idx, s1Name, s2Name)
        
        sheet1 = wb1.get_sheet_by_name(s1Name)
        sheet2 = wb2.get_sheet_by_name(s2Name)
        sheet3 = wb3.create_sheet(s1Name, sh_idx)
        
        maxRowRng = getRangeWarpper(sheet1.max_row, sheet2.max_row)
        maxColRng = getRangeWarpper(sheet1.max_column, sheet2.max_column)
        
        for row_idx in range(1, 43) : #maxRowRng.left <-- RCRP0C210_C8_1 最多到 row:43
            row_idx_1 = row_idx
            row_idx_2 = row_idx
            
            diff = _ColDiff.findRowFrom(row_idx)
            if diff :
                row_idx_2 = diff.rowTo
                    
            row1 = excelUtil.getRows(sheet1)[row_idx_1]
            row2 = excelUtil.getRows(sheet2)[row_idx_2]
            
            for col_idx in range(1, maxColRng.left) :
#                 print("--------------------------------------------")
#                 print(excelUtil.getCellDefine(row1[col_idx -1]))
#                 print(excelUtil.getCellDefine(row2[col_idx -1]))
                col_idx_1 = col_idx
                col_idx_2 = col_idx

                if diff :
                    col_idx_2 = col_idx + diff.colDiff
                
                cell1Val = excelUtil.getCellValue(col_idx_1, row1)
                cell2Val = excelUtil.getCellValue(col_idx_2, row2)
                cell1Val = fixCellValue_None(cell1Val)
                cell2Val = fixCellValue_None(cell2Val)
                cell1Val = stringUtil.trimSpace(cell1Val)
                cell2Val = stringUtil.trimSpace(cell2Val)
                cell1Val = fixCellValue_Zero(cell1Val)
                cell2Val = fixCellValue_Zero(cell2Val)
                cell1Val = fixCellValue_Replace(cell1Val)
                cell2Val = fixCellValue_Replace(cell2Val)
                
                colEnglish = excelUtil.cellEnglishToPos_toStr(col_idx)
    
                newCell = sheet3[colEnglish + str(row_idx)]
                
                if not isNumberEqual(cell1Val, cell2Val) and cell1Val != cell2Val:
                    newCell.value = cell1Val + "," + cell2Val
                    excelUtil.setCellColor(newCell, bgColor="C7EDCC")
                    
#                     print("diff ", s1Name, colEnglish + str(row_idx), "c1=", cell1Val, "c2=", cell2Val)
                else :
                    newCell.value = cell1Val
                    pass
        
#         break
                    
    wb3.save(fileUtil.getDesktopDir() + os.sep + "sample.xlsx")
                
#         openpyxl.worksheet.worksheet.Worksheet

class _ColDiff(Enum):
    B8 = ("B8",8, 8, 0)
    B9 = ("B9",9, 9, 0)
    B13 = ("B13",13, 13, 0)
    B14 = ("B14",14, 14, 0)
    B18 = ("B18",18, 18, 0)
    B19 = ("B19",19, 19, 0)
    B23 = ("B23",23, 23, 0)
    B24 = ("B24",24, 24, 0)
    C25 = ("B25",25, 25, -1)
    C26 = ("B26",26, 26, -1)
    C27 = ("B28",27, 28, -1)
    C28 = ("B29",28, 29, -1)
    C29 = ("B30",29, 30, -1)
    C30 = ("B31",30, 31, -1)
    C31 = ("B33",31, 33, -1)
    C32 = ("B34",32, 34, -1)
    C33 = ("B35",33, 35, -1)
    C34 = ("B36",34, 36, -1)
    C35 = ("B37",35, 37, -1)
    C36 = ("B38",36, 38, -1)
    C37 = ("B39",37, 39, -1)
    C38 = ("C41",38, 41, 0)
    C39 = ("C42",39, 42, 0)
    C40 = ("C43",40, 43, 0)
    C41 = ("C44",41, 44, 0)
    C42 = ("C45",42, 45, 0)
    C43 = ("C46",43, 46, 0)
    
    def __init__(self, startLabel, rowFrom, rowTo, colDiff):
        self.startLabel = startLabel
        self.rowFrom = rowFrom
        self.rowTo = rowTo
        self.colDiff = colDiff
        
    @classmethod
    def findRowFrom(clz, rowFrom):  
        for i, name in enumerate(_ColDiff.__members__, 0):
            e = _ColDiff[name]
            if e.rowFrom == rowFrom :
                return e
        return None    
    

if __name__ == '__main__':
    file1 = fileUtil.getDesktopDir() + os.sep + "RCRP0C210_C8_1.XLSX"
    file2 = fileUtil.getDesktopDir() + os.sep + "RCRP0C210_I8_1.XLSX"
    mainDetail(file1, file2)
    print("done...")
