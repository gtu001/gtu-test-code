import requests
from bs4 import BeautifulSoup as bs
from selenium import webdriver
from gtu.reflect import checkSelf

from selenium.webdriver.remote.webdriver import WebDriver as WD
from gtu.web_crawler.selenium_test import seleniumUtil
from threading import Thread
from gtu.net import ssl_util

from gtu.util import queue_test_001
from gtu.string import stringUtil

import sys
import time
import re
from enum import Enum
from gtu.io import fileUtil
from gtu.error import errorHandler

from abc import ABCMeta, abstractmethod

from gtu.io import LogWriter
from selenium.webdriver.support.ui import WebDriverWait
from gtu.openpyxl_test import excelUtil
from gtu.datetime import dateUtil
from openpyxl import Workbook
from gtu.openpyxl_test import excelUtil


log = LogWriter.LogWriter()


class FirstBankHandler(Thread, metaclass=ABCMeta) :
    loginURL = "https://ccard.firstbank.com.tw/cmsweb/home/index"
    workbookPath = fileUtil.getDesktopDir("firstBankCreditCardCheck_" + dateUtil.currentDatetime_str() + ".xlsx")
    monthIndex = -1
    Excel = None

    def __init__(self) :
        super().__init__()
        self.queue = queue_test_001.DequeObj()
        self.driver = seleniumUtil.getDriver()
        self.start()

    def reflashName(self) :
        newName = "".format(\
            ""
            )
        # self.setName(newName)


    def getWorkbook(self) :
        FirstBankHandler.Excel = excelUtil.WorkbookHolder(FirstBankHandler.workbookPath)
        FirstBankHandler.Excel.loadWorkbook(create=True)
        return FirstBankHandler.Excel.getWorkbook()


    def saveWorkbook(self) :
        FirstBankHandler.Excel.saveWorkbook()


    def findButton(self, text) :
        findBtns = self.driver.find_elements_by_css_selector("button")
        for i,v in enumerate(findBtns) :
            if text in v.text :
                return v
        return None


    def login(self) :
        self.driver.get(FirstBankHandler.loginURL)

        txtIDCard = self.driver.find_element_by_css_selector("input[id=txtIDCard]")
        txtUserAcct = self.driver.find_element_by_css_selector("input[id=txtUserAcct]")
        txtPswd = self.driver.find_element_by_css_selector("input[id=txtPswd]")
        _txtPicCode_myPicVerify = self.driver.find_element_by_css_selector("input[id=_txtPicCode_myPicVerify]")

        loginBtn = self.findButton('登入')
        
        seleniumUtil.WebElementControl.setValue(txtIDCard, "E123474736")
        seleniumUtil.WebElementControl.setValue(txtUserAcct, "gtu001")

        def waitUntilFunc(driver) :
            print("waitUntilFunc try...")
            chk1 = stringUtil.isNotBlank(seleniumUtil.WebElementControl.getValue(txtPswd))
            chk2 = len(seleniumUtil.WebElementControl.getValue(_txtPicCode_myPicVerify)) == 4
            print("check11 = " , chk1 , " , ", chk2)
            print("check22 = " , seleniumUtil.WebElementControl.getValue(txtPswd) , " , ", seleniumUtil.WebElementControl.getValue(_txtPicCode_myPicVerify))
            if chk1 and chk2 :
                return True
            return False
        
        seleniumUtil.WebElementControl.until(self.driver, timeout=300, frequence=0.5, message='', waitUntilFunc=waitUntilFunc) 
        print("loginBtn = ", loginBtn)
        seleniumUtil.WebElementControl.click(loginBtn)


    def waitSelectOptionsFull(self) :
        def waitUntilFunc(driver) :
            options = driver.find_elements_by_xpath("//select[contains(@id, 'selBillDate')]/option")
            print("check_options_length = " , len(options))
            if len(options) == 12 :
                return True
            return False        
        seleniumUtil.WebElementControl.until(self.driver, timeout=300, frequence=0.5, message='', waitUntilFunc=waitUntilFunc)


    def choiceMonthSelect(self) :
        '''
            選擇月份
        '''
        select = seleniumUtil.WebElementControl.waitPageElementByCss("select[id=selBillDate]", '', self.driver)
        self.waitSelectOptionsFull()
        options = seleniumUtil.WebElementControl.getOptionsLst(select)
        for op in range(0, len(options)) :
            seleniumUtil.WebElementControl.setSelect(select, index=op)

        FirstBankHandler.monthIndex += 1
        if FirstBankHandler.monthIndex == 12 :
            return None
        sheetName = options[FirstBankHandler.monthIndex][1]
        print("select date = ", sheetName)

        seleniumUtil.WebElementControl.setSelect(select, index=FirstBankHandler.monthIndex)        
        seleniumUtil.WebElementControl.clickUntil(self.driver, xpath="//button[text()='查詢']")

        dateObj = dateUtil.taiwanDateToDateObj(sheetName, delimit='/')
        dateStr2 = dateUtil.formatDatetimeByJavaFormat(dateObj, 'yyyy/MM/dd')
        print("wait date = ", dateStr2)

        # 等畫面出現正確日其
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support import expected_conditions as EC
        WebDriverWait(self.driver, 30).until(
            EC.text_to_be_present_in_element(
                (By.ID, 'tbQry1'), dateStr2))
        return sheetName



    def showBillPage(self) :
        Alink = seleniumUtil.WebElementControl.waitPageElementByCss("li[class=cms-bg-menu]", '信用卡帳務服務', self.driver)
        seleniumUtil.WebElementControl.click(Alink)
        Blink = seleniumUtil.WebElementControl.waitPageElementByCss("li > a > div", '帳單明細查詢╱列印繳款聯', self.driver)
        seleniumUtil.WebElementControl.click(Blink)

        while True:
            #選擇月份
            sheetName = self.choiceMonthSelect()

            if sheetName == None :
                print("------All Done !!")
                break

            Clink = seleniumUtil.WebElementControl.waitPageElementByCss("li[class=active] > a", '台幣', self.driver)
            
            pageContent = self.driver.page_source
            print("################# start ")

            wb = self.getWorkbook()

            for again in range(0, 10) :
                from selenium.common.exceptions import StaleElementReferenceException
                try:
                    ws = excelUtil.createSheet(sheetName, wb)
                    ws.append(['','消費日','入帳日','消費明細','台幣入帳金額','外幣折算日及金額'])

                    trs = self.driver.find_elements_by_xpath("//div[contains(@class,'cms-label')][text()='信用卡消費明細']/following-sibling::table//tbody//tr")
                    for i,tr in enumerate(trs) :
                        print(i, tr.text)
                        tds = tr.find_elements_by_xpath(".//td")
                        print("--------------")
                        lst = list()
                        for j,td in enumerate(tds) :
                            print(j, td.text)
                            lst.append(td.text)
                        ws.append(lst)
                        
                    self.saveWorkbook()
                    break
                except StaleElementReferenceException as ex :
                    time.sleep(3)
        

         
    def run(self) :
        # bs(self.driver.page_source, "html.parser")
        # while True :
        #     if self.queue.length() > 0 :
        #         sourceStr = self.queue.popright()
        #         _sourceStr = self.beforeProcess(sourceStr)
        #         resultStr = self.searchInput(_sourceStr)
        #         self.myProcess(sourceStr, resultStr)
        #     self.reflashName()
        #     time.sleep(0.2)
        pass




THREAD = FirstBankHandler()




def main() :
    # lst = fileUtil.loadFile_asList(fileUtil.getDesktopDir(fileName="GoogleWord.txt"))
    # for i,v in enumerate(lst) :
    #     THREAD.queue.appendleft(v)
    # pass
    
    THREAD.login()
    THREAD.showBillPage()


if __name__ == '__main__' :
    main()

    # network("http://www.gamer.com.tw") # http://www.91porn.com/view_video.php?viewkey=a9a7ceb02bab655e0e6a

